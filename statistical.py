
import pandas as pd
import os
from datetime import datetime
from copy import deepcopy
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import shutil
import numpy as np
import traceback
import logging

# time_ymd='20200513'  
time_ymd=datetime.strftime(datetime.now(), '%Y%m%d')
time_ymdHMS = datetime.strftime(datetime.now(), '%Y%m%d %H:%M:%S')
time_hm = datetime.strftime(datetime.now(), '%H:%M')
logging.basicConfig(filename='error.txt', level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')
errorFile = open('./error.txt', 'a')

class WorkRecode():
     
    # def is_datetime(self, dt):
    #     if str(dt) == time_ymd:
    #         return 1
    #     else:
    #         return 0

    def merge(self, config_wb, config_ws): 
        if not os.path.exists('sum.xlsx'):
            shutil.copy('./workbook.xlsx','./sum.xlsx')

        #读取文件夹中的xlsx文件
        dflist = []
        print('正在收集文件')
        for file in os.listdir('./recode/'):
            try:
                filename = os.path.join('./recode/'+file)
                df = pd.read_excel(filename, sheet_name='Sheet1', skiprows=([0,1]))
                dflist.append(df)
            except Exception as e:    
                errorFile.write(traceback.format_exc())
                errorFile.close()
                continue

        #存成读取总记录文件
        sum_wb = load_workbook('./sum.xlsx')
        sum_ws = sum_wb['Sheet1']
        logs_ws = config_wb['logs']
        # sum_ws_logs = sum_wb['logs']
        # sum_df = pd.read_excel('sum.xlsx')
        #读取sum_ws的总行 总列
        rows = sum_ws.max_row
        cols = sum_ws.max_column
        # print(rows)
        time_col_data = []
        for i in range(4, rows+1):
            cell_val = sum_ws.cell(row = i, column = 2).value
            time_col_data.append(str(cell_val))
        if time_ymd in time_col_data:
            for i in range(4, rows+1):
                # print(sum_ws.cell(row = i, column = 4).value)
                if str(sum_ws.cell(row = i, column = 2).value) == time_ymd:
                    sum_ws.delete_rows(i, i+len(time_col_data))
            sum_wb.save('./sum.xlsx')
            sum_wb.close()

        sum_wb = load_workbook('./sum.xlsx')
        sum_ws = sum_wb['Sheet1']

        flag = 0
        for df in dflist:
            try:
            #文件没填写不插入
                if df.shape[0] == 0:
                    logs_ws.append(['用户%s未填写' % os.listdir('./recode/')[flag].split('.')[0].split('_')[0], time_ymdHMS])
                    config_wb.save('./config.xlsx')
                    continue
                flag_insert = 0
                for i in range(df.shape[0]):
                    if str((df.iloc[i])[1]) == time_ymd:
                        sum_ws.append(list(df.iloc[i]))
                        print('插入成功', i)
                        flag_insert = 1
                
                        
                if flag_insert == 0:
                    logs_ws.append(['用户%s未填写' % os.listdir('./recode/')[flag].split('.')[0].split('_')[0], time_ymdHMS])
                
                
            except Exception as e:
                print(e)
                errorFile.write(traceback.format_exc())
                errorFile.close()
                continue
            flag = flag+1
        sum_wb.save('./sum.xlsx')
        sum_wb.close()
        config_ws.cell(2, 2).value = time_ymd
        print('收集完成')
    
    def distribute(self, config_wb, config_ws):
        workbook_name = config_ws.cell(2, 4).value
        #判断这个字符汉字
        def is_chinese(chr):
            if chr >= '\u4e00' and chr <= '\u9fa5':
                return 0
            else:
                return 1

        #获取当前日期
        now_time = datetime.strftime(datetime.now(), '%Y%m%d')
        #读取名单
        name_df = pd.read_excel('./config.xlsx', sheet_name='names')
        #按名单创建 姓名-日期 格式的xlsx文件

        #创建文件夹存放拆分好的excel表
        if not os.path.exists('./recode'):
            os.mkdir('./recode')
        else:
            shutil.rmtree('./recode')
            os.mkdir('./recode')


        for name in name_df['姓名']:
        # 判断 名字不为空 且都是汉字
            try:    
                if name != ' ' and sum([is_chinese(i) for i in str(name)]) == 0:
                    excel_name = str(name)+'-'+now_time
                    shutil.copy('./'+workbook_name,'./recode/'+excel_name+'.xlsx')
            except Exception as e:
                print(e)
                errorFile.write(traceback.format_exc())
                errorFile.close()
                continue
        config_ws.cell(2, 5).value = 1
        print('分发完成')





if __name__ == '__main__':
    work_recode = WorkRecode()
    config_wb = load_workbook('./config.xlsx')
    config_ws = config_wb['config']
    if config_ws.cell(2, 5).value == 0:
        if time_hm > str(config_ws.cell(2, 3).value):
            work_recode.distribute(config_wb, config_ws)
        else:
            print('时间未到')
    else:
        print('已经完成分发')

    if str(config_ws.cell(2, 2).value) != time_ymd:
        if time_hm > str(config_ws.cell(2, 1).value):
            work_recode.merge(config_wb, config_ws)
        else:
            print('时间未到')
    else:
        print('今日已经收集')
    config_wb.save('./config.xlsx')
    config_wb.close()